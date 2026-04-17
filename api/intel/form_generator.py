"""PDS XLSX generation (hackathon path — Prompt B.6 / BE13).

Template path: ``api/templates/forms/csc_form_212.xlsx`` (under the ``api`` package;
plural ``templates``, not ``template``).
"""

from __future__ import annotations

import asyncio
import logging
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from api.intel.form_mappings.csc_form_212 import (
    CSC_FORM_212_CELL_MAPPING,
    _format_value,
    resolve_source,
    validate_enum_value,
)
from api.runtime_paths import ensure_output_dir

logger = logging.getLogger(__name__)

PDS_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1] / "templates" / "forms" / "csc_form_212.xlsx"
)
OUTPUT_DIR = str(ensure_output_dir())

# CSC Form 212 Sheet C1 — cells bound to enum-validated profile fields
_CELL_ENUM_FIELDS = {"D16": "sex", "D17": "civil_status"}


def _generate_pds_xlsx_sync(registration_id: int, context: dict[str, Any]) -> str:
    """Blocking: copy template, fill cells, save. Raises ``FileNotFoundError`` if template missing."""
    if not PDS_TEMPLATE_PATH.is_file():
        raise FileNotFoundError(
            f"PDS template not found (expected file): {PDS_TEMPLATE_PATH}"
        )

    output_path = Path(OUTPUT_DIR) / f"{registration_id}_pds.xlsx"
    shutil.copy(PDS_TEMPLATE_PATH, output_path)

    wb = load_workbook(output_path)
    try:
        for sheet_name, cells in CSC_FORM_212_CELL_MAPPING.items():
            if sheet_name not in wb.sheetnames:
                logger.warning(
                    "PDS workbook missing sheet %r; skipping mapped cells for it",
                    sheet_name,
                )
                continue
            ws = wb[sheet_name]
            for cell_addr, source_spec in cells.items():
                try:
                    raw = resolve_source(source_spec, context)
                    if cell_addr in _CELL_ENUM_FIELDS and raw is not None:
                        validate_enum_value(
                            _CELL_ENUM_FIELDS[cell_addr], str(raw)
                        )
                    formatted = _format_value(raw)
                    ws[cell_addr].value = formatted
                except Exception:
                    logger.exception(
                        "PDS cell %s!%s (source=%r) failed; leaving cell unchanged",
                        sheet_name,
                        cell_addr,
                        source_spec,
                    )
        wb.save(output_path)
    finally:
        wb.close()

    return str(output_path.resolve())


async def generate_pds_xlsx(registration_id: int, context: dict[str, Any]) -> str:
    """Copy the PDS template, fill mapped cells from *context*, save, return path."""
    return await asyncio.to_thread(
        _generate_pds_xlsx_sync, registration_id, context
    )

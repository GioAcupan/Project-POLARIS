"""CSC Form 212 (PDS) cell mapping for Sheet C1, Page 1.

Run ``audit_mapping_against_template(path)`` against the real
``csc_form_212.xlsx`` before turning off PITCH_MODE in production; empty
result means all mapped cells are writable (sheet exists, merge anchors OK).
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

CIVIL_STATUS_VALUES = ("Single", "Married", "Separated", "Widowed")
SEX_VALUES = ("Male", "Female")

# Sheet C1 — discovered mapping v3.4 §5.4.2
CSC_FORM_212_CELL_MAPPING: dict[str, dict[str, str]] = {
    "C1": {
        "D10": "teacher.last_name",
        "D11": "teacher.first_name",
        "O11": "profile.name_extension",
        "D12": "teacher.middle_name",
        "D13": "profile.date_of_birth",
        "D16": "profile.sex",
        "D17": "profile.civil_status",
        "D34": "teacher.deped_id",
        "I26": "profile.addr_house_no",
        "L26": "profile.addr_street",
        "I27": "profile.addr_subdivision",
        "L27": "profile.addr_barangay",
        "I30": "profile.addr_city",
        "L30": "profile.addr_province",
        "I31": "profile.addr_zip",
        "I32": "profile.telephone_number",
        "I33": "profile.mobile_number",
        "I34": "profile.email",
    },
}


def validate_enum_value(field: str, value: str) -> None:
    """Raise ValueError if *value* is non-empty and not allowed for *field*."""
    if value is None or not str(value).strip():
        return
    v = str(value).strip()
    if field == "sex":
        if v not in SEX_VALUES:
            raise ValueError(f"sex must be one of {SEX_VALUES!r}, got {v!r}")
    elif field == "civil_status":
        if v not in CIVIL_STATUS_VALUES:
            raise ValueError(
                f"civil_status must be one of {CIVIL_STATUS_VALUES!r}, got {v!r}"
            )
    else:
        raise ValueError(f"Unknown enum field: {field!r}")


def resolve_source(source: str, context: dict[str, Any]) -> Any:
    """Resolve a *source* string from *context*.

    - ``constant:VALUE`` — literal ``VALUE`` (only the first ``constant:`` prefix
      is stripped; any further colons remain part of the value).
    - Dotted paths like ``teacher.deped_id`` or ``profile.email`` — walk
      objects or dicts. Root ``profile`` uses ``profile`` or ``profile_extended``
      from *context* (router passes ``profile_extended``).
    """
    prefix = "constant:"
    if source.startswith(prefix):
        return source[len(prefix) :]

    parts = source.split(".")
    if not parts:
        return None

    root = parts[0]
    if root == "profile":
        current: Any = context.get("profile") or context.get("profile_extended")
    else:
        current = context.get(root)

    for part in parts[1:]:
        if current is None:
            return None
        if isinstance(current, dict):
            current = current.get(part)
        else:
            current = getattr(current, part, None)
    return current


def _format_value(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        val = val.date()
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, bool):
        return str(val)
    if isinstance(val, float):
        s = f"{val:.12g}"
        return s
    if isinstance(val, Decimal):
        if val == val.to_integral():
            return str(int(val))
        return format(val, "f").rstrip("0").rstrip(".")
    return str(val)


def _cell_row_col(cell_addr: str) -> tuple[int, int]:
    col_letter, row = coordinate_from_string(cell_addr)
    col_idx = column_index_from_string(col_letter)
    return row, col_idx


def _merge_status_for_cell(
    sheet: Any, row: int, col_idx: int
) -> tuple[bool, bool]:
    """Return (is_in_merged_range, is_top_left_anchor)."""
    for mr in sheet.merged_cells.ranges:
        if (
            mr.min_row <= row <= mr.max_row
            and mr.min_col <= col_idx <= mr.max_col
        ):
            anchor = row == mr.min_row and col_idx == mr.min_col
            return True, anchor
    return False, False


def audit_mapping_against_template(path: str) -> dict[str, str]:
    """Check each mapped cell exists and is writable (unmerged or merge anchor).

    Returns warnings keyed by ``f\"{sheet}!{cell}\"``; empty dict means clean.
    """
    wb = load_workbook(path, read_only=False, data_only=False)
    warnings: dict[str, str] = {}
    try:
        for sheet_name, cells in CSC_FORM_212_CELL_MAPPING.items():
            if sheet_name not in wb.sheetnames:
                for cell_addr in cells:
                    key = f"{sheet_name}!{cell_addr}"
                    warnings[key] = f"sheet {sheet_name!r} not in workbook"
                continue
            ws = wb[sheet_name]
            for cell_addr in cells:
                key = f"{sheet_name}!{cell_addr}"
                try:
                    row, col_idx = _cell_row_col(cell_addr)
                except Exception as e:
                    warnings[key] = f"invalid cell address: {e}"
                    continue
                in_merge, is_anchor = _merge_status_for_cell(ws, row, col_idx)
                if in_merge and not is_anchor:
                    warnings[key] = (
                        "cell lies in merged range but is not the top-left anchor"
                    )
    finally:
        wb.close()
    return warnings
